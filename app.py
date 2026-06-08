# ╔══════════════════════════════════════════════════════════════════╗
# ║  투찰전략 분석 시스템 v2.4                                      ║
# ║  개선: 패턴 범위 비교 강화                                     ║
# ║  - ①패턴: 현재까지의 낙찰이력 차트 기반 분석                   ║
# ║  - 한전 전체 / 동일 발주처 / 한전 감리·진단 / 동일 발주처 분야  ║
# ║  - ③트렌드 최소값 보정 (±0.02% 미만 시 보정)                  ║
# ║  - ②유사표본: 용역성격/지역/업체수구간 + 직전5건 최소거리       ║
# ╚══════════════════════════════════════════════════════════════════╝

import streamlit as st
import pandas as pd
import numpy as np
import xlrd, io, os, json
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from datetime import datetime

plt.rcParams.update({"font.family": "DejaVu Sans", "axes.unicode_minus": False})

st.set_page_config(page_title="투찰전략 분석 시스템", page_icon="📊", layout="wide")
st.markdown("""
<style>
.main-header{background:linear-gradient(135deg,#1a2744,#243260);color:white;
    padding:20px 30px;border-radius:10px;margin-bottom:20px}
.val-box{border-radius:8px;padding:10px 15px;font-weight:bold;
    font-size:1.1em;text-align:center;margin:4px 0}
.val-pattern{background:#dbeafe;color:#1d4ed8}
.val-similar{background:#dcfce7;color:#15803d}
.val-trend{background:#fef9c3;color:#854d0e}
.val-rec{background:#f3e8ff;color:#7c3aed}
.grade-a{background:#dcfce7;color:#15803d;padding:2px 8px;border-radius:4px;font-weight:bold}
.grade-b{background:#dbeafe;color:#1d4ed8;padding:2px 8px;border-radius:4px;font-weight:bold}
.grade-c{background:#fef9c3;color:#854d0e;padding:2px 8px;border-radius:4px;font-weight:bold}
.grade-d{background:#fee2e2;color:#991b1b;padding:2px 8px;border-radius:4px;font-weight:bold}
</style>""", unsafe_allow_html=True)

DATA_DIR     = "data"
HISTORY_FILE = os.path.join(DATA_DIR, "history.pkl")
PATTERN_FILE = os.path.join(DATA_DIR, "pattern_stats.json")
os.makedirs(DATA_DIR, exist_ok=True)

# ── 기초금액 구간별 보정값 ─────────────────────────────────────
AMT_BRACKETS = {
    "~0.5억":  {"range":(0,   0.5),  "adj":+0.0013,"note":"소형"},
    "0.5~1억": {"range":(0.5, 1.0),  "adj":+0.0045,"note":"소형"},
    "1~2억":   {"range":(1.0, 2.0),  "adj":+0.0424,"note":"유리 ↑"},
    "2~5억":   {"range":(2.0, 5.0),  "adj":-0.0295,"note":"보수적 ↓"},
    "5~10억":  {"range":(5.0, 10.0), "adj":+0.0398,"note":"대형"},
    "10억+":   {"range":(10.0,9999), "adj":+0.0373,"note":"대형"},
}

def get_amt_info(base_억):
    for label, info in AMT_BRACKETS.items():
        lo, hi = info["range"]
        if lo <= base_억 < hi:
            return label, info["adj"], info["note"]
    return "미정", 0.0, ""

# ── 낙찰이력 로드 ─────────────────────────────────────────────
@st.cache_data
def load_history():
    if os.path.exists(HISTORY_FILE):
        return pd.read_pickle(HISTORY_FILE)
    return None

@st.cache_data
def load_pattern_stats():
    if os.path.exists(PATTERN_FILE):
        with open(PATTERN_FILE, encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_history(df):
    os.makedirs(DATA_DIR, exist_ok=True)
    df.to_pickle(HISTORY_FILE)
    st.cache_data.clear()

def save_pattern_stats(stats):
    with open(PATTERN_FILE, "w", encoding="utf-8") as f:
        json.dump(stats, f, ensure_ascii=False, indent=2)
    st.cache_data.clear()

# ── 분석 함수 ─────────────────────────────────────────────────
def analyze_pattern(org, df_c, pattern_stats):
    if org in pattern_stats:
        st_d = pattern_stats[org]
        return {
            "pred":         st_d.get("pred", 0),
            "conservative": st_d.get("conservative", 0),
            "aggressive":   st_d.get("aggressive", 0),
            "trend":        st_d.get("trend", "→횡보"),
            "pattern":      st_d.get("pattern", "무작위패턴").replace("패턴",""),
            "autocorr":     st_d.get("autocorr", 0),
            "last_val":     st_d.get("last_val", 0),
            "r5":           st_d.get("r5", 0),
            "r10":          st_d.get("r10", 0),
            "mean":         st_d.get("mean", 0),
            "std":          st_d.get("std", 0.5),
            "n":            st_d.get("n", 0),
            "w5":           st_d.get("w5", 0.25),
            "w10":          st_d.get("w10", 0.20),
            "wm":           st_d.get("wm", 0.55),
            "grade":        st_d.get("grade", "C"),
            "mae":          st_d.get("mae", 0.5),
            "drift6m":      st_d.get("drift6m", 0),
            "trend_boost":  st_d.get("trend_boost", 0),
            "drift_note":   st_d.get("drift_note", ""),
            "recent10":     st_d.get("recent10", []),
            "all_vals":     None,
            "source":       "패턴통계DB"
        }
    if df_c is None: return None
    sub = df_c[df_c["발주기관"]==org]["예가/기초(0%)"].values
    if len(sub) < 5: return None
    n=len(sub); mean=np.mean(sub); std=np.std(sub)
    r5=np.mean(sub[-5:]); r10=np.mean(sub[-10:]) if n>=10 else mean
    ac=float(np.corrcoef(sub[:-1],sub[1:])[0,1]) if n>=3 else 0
    coef=float(np.polyfit(np.arange(min(20,n)),sub[-min(20,n):],1)[0])
    trend   = "↑상승" if coef>0.02 else "↓하락" if coef<-0.02 else "→횡보"
    pattern = "연속성" if ac>0.2 else "반전" if ac<-0.2 else "무작위"
    w5,w10,wm = 0.25,0.20,0.55
    pred = w5*r5 + w10*r10 + wm*mean
    lv = float(sub[-1])
    adj = (lv*abs(ac)*0.2 if pattern=="연속성" else
           -lv*abs(ac)*0.3 if pattern=="반전" else 0.0)
    pred_final = round(pred+adj, 4)
    errs=[abs((0.25*np.mean(sub[:i][-5:])+0.20*(np.mean(sub[:i][-10:]) if i>=10 else np.mean(sub[:i]))+0.55*np.mean(sub[:i]))-sub[i])
          for i in range(min(10,n//2),n)]
    mae=np.mean(errs) if errs else 0.5
    grade="A" if mae<0.35 else "B" if mae<0.45 else "C" if mae<0.55 else "D"
    return {
        "pred":pred_final,"conservative":round(pred_final-std*0.4,4),
        "aggressive":round(pred_final+std*0.4,4),
        "trend":trend,"pattern":pattern,"autocorr":round(ac,4),
        "last_val":round(lv,4),"r5":round(r5,4),"r10":round(r10,4),
        "mean":round(mean,4),"std":round(std,4),"n":n,
        "w5":w5,"w10":w10,"wm":wm,"grade":grade,"mae":round(mae,4),
        "drift6m":0.0,"trend_boost":0.0,"drift_note":"",
        "recent10":[round(float(v),4) for v in sub[-10:]],
        "all_vals":sub.tolist(),"source":"직접계산"
    }

def _service_keywords(name):
    name = str(name)
    groups = [
        ("광학", ["광학"]),
        ("VLF", ["VLF"]),
        ("PD", ["PD"]),
        ("콘크리트", ["콘크리트"]),
        ("초음파", ["초음파"]),
        ("감리", ["감리"]),
        ("진단", ["진단"] + DIAG_KWS),
        ("설계", ["설계"]),
        ("측정", ["측정"]),
    ]
    labels, kws = [], []
    for label, words in groups:
        if any(w in name for w in words):
            labels.append(label)
            kws.extend(words)
    if not kws:
        labels, kws = ["감리/진단"], ["감리"] + DIAG_KWS
    return "+".join(dict.fromkeys(labels)), list(dict.fromkeys(kws))

def _region_key(text):
    text = str(text or "")
    for key in ["서울","부산","대구","인천","광주","대전","울산","세종","경기","강원","충북","충남","전북","전남","경북","경남","제주"]:
        if key in text:
            return key
    return ""

def _company_bucket(v):
    if pd.isna(v):
        return None
    try:
        n = int(float(v))
    except (TypeError, ValueError):
        return None
    if n < 0:
        return None
    lo = (n // 10) * 10
    return lo, lo + 9, f"{lo}~{lo+9}개"

def _sort_by_open_date(df):
    if df is None or len(df)==0:
        return df
    out = df.copy()
    if "개찰일" in out.columns:
        parsed = pd.to_datetime(out["개찰일"].astype(str), errors="coerce")
        short = parsed.isna() & out["개찰일"].astype(str).str.match(r"^\d{2}\.\d{1,2}\.\d{1,2}$", na=False)
        if short.any():
            parsed.loc[short] = pd.to_datetime("20" + out.loc[short, "개찰일"].astype(str), format="%Y.%m.%d", errors="coerce")
        out = out.assign(_sort_date=parsed).sort_values(["_sort_date"], na_position="first")
    return out

def _similar_pattern_prediction(sim, window=5, top_k=5):
    vals = sim["예가/기초(0%)"].dropna().values if "예가/기초(0%)" in sim.columns else np.array([])
    match = _nearest_trend_prediction(vals, window=window, top_k=top_k)
    if match:
        return {
            "pred": match["pred"],
            "method": f"직전{window}건 유사패턴",
            "matched_n": len(match["matches"]),
            "best_distance": float(match["best_distance"]),
            "next_vals": [round(float(v),4) for v in match["next_vals"]],
        }
    if len(vals)==0:
        return None
    weights = np.linspace(0.5, 1.5, len(vals))
    return {
        "pred": float(np.average(vals, weights=weights)),
        "method": "유사표본 가중평균",
        "matched_n": 0,
        "best_distance": None,
        "next_vals": [],
    }

def _similar_candidate_from_pool(label, pool, window=5, basis=""):
    pool = _sort_by_open_date(pool)
    if pool is None or len(pool) < 3:
        return None
    pred_info = _similar_pattern_prediction(pool, window=window, top_k=5)
    if not pred_info:
        return None
    vals = pool["예가/기초(0%)"].dropna().values
    if len(vals) < 3:
        return None
    co = pool["업체수"].mean() if "업체수" in pool.columns else None
    return {
        "label": label,
        "pred": round(float(pred_info["pred"]), 4),
        "n": len(vals),
        "mean": round(float(np.mean(vals)), 4),
        "std": round(float(np.std(vals)), 4),
        "avg_companies": round(float(co), 1) if co is not None and not np.isnan(float(co)) else None,
        "method": pred_info["method"],
        "best_distance": round(float(pred_info["best_distance"]), 4) if pred_info["best_distance"] is not None else None,
        "matched_n": pred_info["matched_n"],
        "next_vals": pred_info["next_vals"],
        "basis": basis,
    }

def _pick_best_candidate(candidates):
    valid = [c for c in candidates if c]
    if not valid:
        return None
    with_dist = [c for c in valid if c.get("best_distance") is not None]
    if with_dist:
        return sorted(with_dist, key=lambda c: (c["best_distance"], -c["n"], c["std"]))[0]
    return sorted(valid, key=lambda c: (c["std"], -c["n"]))[0]

def analyze_similar(name, base_원, df_c, region=""):
    """② 유사표본: 용역성격/지역/업체수 10단위 구간으로 표본화 후 직전 5건 유사패턴 분석."""
    if df_c is None or base_원<=0: return None
    label, kws = _service_keywords(name)
    work = df_c.copy()
    mask = pd.Series([False]*len(work), index=work.index)
    for kw in kws:
        mask = mask | work["공고명"].astype(str).str.contains(kw, na=False, regex=False)
    amt_mask = (work["기초금액"]>=base_원*0.5) & (work["기초금액"]<=base_원*1.5)
    base_pool = work[mask & amt_mask].copy()
    if len(base_pool) < 7:
        amt_mask = (work["기초금액"]>=base_원*0.3) & (work["기초금액"]<=base_원*2.0)
        base_pool = work[mask & amt_mask].copy()

    region_label = _region_key(region) or _region_key(name)
    candidates = []
    field_pool = base_pool.copy()
    candidates.append(_similar_candidate_from_pool(
        f"분야:{label}", field_pool, window=5,
        basis=f"용역성격 {label}, 기초금액 유사범위"
    ))

    region_pool = pd.DataFrame()
    if region_label and "지역" in base_pool.columns:
        rmask = base_pool["지역"].astype(str).str.contains(region_label, na=False, regex=False)
        region_pool = base_pool[rmask].copy()
        candidates.append(_similar_candidate_from_pool(
            f"지역:{region_label}", region_pool, window=5,
            basis=f"{region_label} 지역 + 용역성격 {label}"
        ))

    target_bucket = None
    bucketed = pd.DataFrame()
    if "업체수" in base_pool.columns and base_pool["업체수"].notna().any():
        avg_for_bucket = base_pool["업체수"].dropna().mean()
        target_bucket = _company_bucket(avg_for_bucket)
        if target_bucket:
            blo, bhi, _ = target_bucket
            bucketed = base_pool[(base_pool["업체수"]>=blo) & (base_pool["업체수"]<=bhi)].copy()
            candidates.append(_similar_candidate_from_pool(
                f"업체수:{target_bucket[2]}", bucketed, window=5,
                basis=f"참여업체수 {target_bucket[2]} 구간"
            ))

    fallback = False
    if not any(candidates):
        if is_diag(name):
            kepco_mask = work['발주기관'].astype(str).str.contains('한국전력공사',na=False,regex=False)
            diag_mask  = pd.Series([False]*len(work), index=work.index)
            for kw in DIAG_KWS:
                diag_mask = diag_mask | work['공고명'].astype(str).str.contains(kw,na=False,regex=False)
            sim = work[kepco_mask & diag_mask].copy()
            label, kws = "진단", DIAG_KWS
            fallback = True
            candidates.append(_similar_candidate_from_pool("분야대체:진단", sim, window=5, basis="한전 진단 전체 표본"))
        elif is_supervision(name):
            kepco_mask = work['발주기관'].astype(str).str.contains('한국전력공사',na=False,regex=False)
            sup_mask   = work['공고명'].astype(str).str.contains('감리',na=False,regex=False)
            sim = work[kepco_mask & sup_mask].copy()
            label, kws = "감리", ["감리"]
            fallback = True
            candidates.append(_similar_candidate_from_pool("분야대체:감리", sim, window=5, basis="한전 감리 전체 표본"))

    candidates = [c for c in candidates if c]
    best = _pick_best_candidate(candidates)
    if not best: return None
    return {
        "pred":best["pred"],
        "n":best["n"],
        "mean":best["mean"],
        "std":best["std"],
        "avg_companies":best.get("avg_companies"),
        "keywords":kws,
        "service_label":label,
        "region_label":region_label if region_label else "",
        "company_bucket":target_bucket[2] if target_bucket else "",
        "method":best["method"],
        "matched_n":best["matched_n"],
        "best_distance":best.get("best_distance"),
        "next_vals":best.get("next_vals",[]),
        "fallback":fallback,
        "fallback_note":"표본 부족으로 한전 분야 표본 대체" if fallback else "",
        "selected_label":best["label"],
        "selected_basis":best.get("basis",""),
        "candidates":candidates,
    }

def _trend_field_filter(name):
    if is_diag(name):
        return "진단", DIAG_KWS
    if is_supervision(name):
        return "감리", ["감리"]
    return "전체", []

def _trend_scope_df(org, name, df_c):
    if df_c is None:
        return pd.DataFrame()
    field_label, keywords = _trend_field_filter(name)
    sub = df_c[df_c["발주기관"] == org].copy()
    if keywords:
        mask = pd.Series([False]*len(sub), index=sub.index)
        for kw in keywords:
            mask = mask | sub["공고명"].astype(str).str.contains(kw, na=False, regex=False)
        sub = sub[mask].copy()
    if len(sub) < 21 and keywords:
        # 동일 발주처 분야 표본이 부족하면 같은 분야의 한전 전체 표본으로 확장한다.
        sub = df_c[df_c["발주기관"].astype(str).str.contains("한국전력공사", na=False, regex=False)].copy()
        mask = pd.Series([False]*len(sub), index=sub.index)
        for kw in keywords:
            mask = mask | sub["공고명"].astype(str).str.contains(kw, na=False, regex=False)
        sub = sub[mask].copy()
    if len(sub) < 21:
        # 그래도 부족하면 동일 발주처 전체 표본으로 복귀한다.
        sub = df_c[df_c["발주기관"] == org].copy()
    if "개찰일" in sub.columns:
        parsed = pd.to_datetime(sub["개찰일"].astype(str), errors="coerce")
        short = parsed.isna() & sub["개찰일"].astype(str).str.match(r"^\d{2}\.\d{1,2}\.\d{1,2}$", na=False)
        if short.any():
            parsed.loc[short] = pd.to_datetime("20" + sub.loc[short, "개찰일"].astype(str), format="%Y.%m.%d", errors="coerce")
        sub = sub.assign(_sort_date=parsed).sort_values(["_sort_date"], na_position="first")
    return sub

def _nearest_trend_prediction(vals, window=10, top_k=5):
    vals = np.array([float(v) for v in vals if not pd.isna(v)], dtype=float)
    if len(vals) < window + 2:
        return None
    target = vals[-window:]
    candidates = []
    # 마지막 target과 겹치지 않는 과거 window -> next 값만 후보로 사용한다.
    for start in range(0, len(vals) - window):
        end = start + window
        if end >= len(vals) - 1:
            break
        seq = vals[start:end]
        next_val = vals[end]
        dist = float(np.mean(np.abs(seq - target)))
        candidates.append((dist, start, next_val, seq))
    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0])
    best = candidates[:min(top_k, len(candidates))]
    dists = np.array([b[0] for b in best], dtype=float)
    next_vals = np.array([b[2] for b in best], dtype=float)
    weights = 1 / (dists + 0.0001)
    pred = float(np.average(next_vals, weights=weights))
    spread = float(np.std(next_vals)) if len(next_vals) > 1 else 0.10
    return {
        "pred": pred,
        "lo": pred - max(0.05, spread * 0.5),
        "hi": pred + max(0.05, spread * 0.5),
        "matches": best,
        "target": target,
        "next_vals": next_vals,
        "best_distance": float(best[0][0]),
    }

def _trend_candidate_from_pool(label, sub, window=10, basis=""):
    sub = _sort_by_open_date(sub)
    vals = sub["예가/기초(0%)"].dropna().values if "예가/기초(0%)" in sub.columns else np.array([])
    if len(vals) < 5: return None

    match = _nearest_trend_prediction(vals, window=window, top_k=5)
    rn = min(window, len(vals))
    recent = vals[-rn:]
    older = vals[:-rn]
    rm = float(np.mean(recent))
    om = float(np.mean(older)) if len(older)>0 else rm
    drift = rm - om
    r3 = vals[-3:] if len(vals)>=3 else vals

    if match:
        raw_pred = match["pred"]
        lo2 = round(match["lo"], 2)
        hi2 = round(match["hi"], 2)
        if lo2 > hi2:
            lo2, hi2 = hi2, lo2
        method = f"직전{window} 유사패턴"
        matched_n = len(match["matches"])
        best_distance = round(match["best_distance"], 4)
        next_vals = [round(float(v), 4) for v in match["next_vals"]]
    else:
        raw_pred = rm + drift*0.3
        if abs(raw_pred) < 0.02:
            raw_pred = float(np.mean(vals))
        lo2 = round(raw_pred - 0.10, 2)
        hi2 = round(raw_pred + 0.10, 2)
        method = "최근흐름 보정"
        matched_n = 0
        best_distance = None
        next_vals = []

    co = sub["업체수"].tail(rn).mean() if "업체수" in sub.columns else None
    return {
        "label":label,
        "pred":round(float(raw_pred),4),
        "recent_mean":round(rm,4),
        "drift":round(drift,4),
        "recent_n":rn,
        "recent3_mean":round(float(np.mean(r3)),4),
        "avg_companies":round(float(co),1) if co and not np.isnan(float(co)) else None,
        "trend_lo":lo2,
        "trend_hi":hi2,
        "trend_range":f"{lo2:+.2f}%~{hi2:+.2f}%",
        "method":method,
        "matched_n":matched_n,
        "best_distance":best_distance,
        "next_vals":next_vals,
        "field_label":label,
        "scope_n":len(vals),
        "n":len(vals),
        "std":round(float(np.std(vals)),4),
        "basis":basis,
    }

def analyze_trend(org, name, df_c, region=""):
    """③ 트렌드: 전체/분야/지역 후보 중 유사패턴 가능성이 높은 값을 기본 추천."""
    if df_c is None: return None
    field_label, keywords = _trend_field_filter(name)
    candidates = []
    overall = df_c[df_c["발주기관"] == org].copy()
    candidates.append(_trend_candidate_from_pool("전체", overall, window=10, basis="동일 발주처 전체 흐름"))

    field_pool = _trend_scope_df(org, name, df_c)
    candidates.append(_trend_candidate_from_pool(f"분야:{field_label}", field_pool, window=10, basis=f"{field_label} 분야 흐름"))

    region_label = _region_key(region) or _region_key(name)
    if region_label and "지역" in df_c.columns:
        region_pool = df_c[df_c["지역"].astype(str).str.contains(region_label, na=False, regex=False)].copy()
        if keywords:
            rmask = pd.Series([False]*len(region_pool), index=region_pool.index)
            for kw in keywords:
                rmask = rmask | region_pool["공고명"].astype(str).str.contains(kw, na=False, regex=False)
            region_pool = region_pool[rmask].copy()
        candidates.append(_trend_candidate_from_pool(f"지역:{region_label}", region_pool, window=10, basis=f"{region_label} 지역 흐름"))

    candidates = [c for c in candidates if c]
    best = _pick_best_candidate(candidates)
    if not best: return None
    best = dict(best)
    best["selected_label"] = best["label"]
    best["selected_basis"] = best.get("basis","")
    best["candidates"] = candidates
    return best

def recommend_range(a1,a2,a3):
    vals=[v["pred"] for v in [a1,a2,a3] if v]
    if not vals: return None,None
    mv=np.mean(vals); sv=np.std(vals) if len(vals)>1 else 0.1
    return round(mv-sv*0.5,4),round(mv+sv*0.5,4)

def convergence_score(a1,a2,a3):
    vals=[v["pred"] for v in [a1,a2,a3] if v]
    if len(vals)<2: return None,"데이터부족"
    sv=np.std(vals)
    if sv<0.05:   return sv,"★★★ 높음"
    elif sv<0.10: return sv,"★★☆ 보통"
    elif sv<0.20: return sv,"★☆☆ 낮음"
    else:         return sv,"⚠️ 분산큼"

def pattern_candidates(a1):
    if not a1:
        return []
    return [
        {
            "label":"보수값",
            "pred":a1.get("conservative", a1["pred"]),
            "basis":f"기준값에서 표준편차 보정. n={a1.get('n',0)} / std={a1.get('std',0):.4f}",
        },
        {
            "label":"기준값(추천)",
            "pred":a1["pred"],
            "basis":f"최근5건 {a1.get('r5',0):+.4f}, 최근10건 {a1.get('r10',0):+.4f}, 전체평균 {a1.get('mean',0):+.4f}",
        },
        {
            "label":"공격값",
            "pred":a1.get("aggressive", a1["pred"]),
            "basis":f"기준값에 표준편차 보정. 패턴={a1.get('pattern','-')} / 추세={a1.get('trend','-')}",
        },
    ]

def _option_text(opt):
    dist = opt.get("best_distance")
    dist_txt = f" | MAE {dist:.4f}" if dist is not None else ""
    n_txt = f" | n={opt.get('n')}" if opt.get("n") is not None else ""
    return f"{opt.get('label','후보')} {opt.get('pred',0):+.4f}%{n_txt}{dist_txt}"

def select_candidate_ui(title, candidates, key, default_label=None):
    candidates = [c for c in candidates if c]
    if not candidates:
        st.caption("선택 가능한 후보가 없습니다.")
        return None
    default_idx = 0
    if default_label:
        for i,opt in enumerate(candidates):
            if opt.get("label") == default_label:
                default_idx = i
                break
    idx = st.radio(
        title,
        options=list(range(len(candidates))),
        index=default_idx,
        format_func=lambda i: _option_text(candidates[i]),
        key=key,
    )
    opt = candidates[idx]
    st.caption(f"근거: {opt.get('basis','-')}")
    return opt

def selected_analysis(base, opt):
    if not base or not opt:
        return base
    out = dict(base)
    out["pred"] = round(float(opt["pred"]), 4)
    out["selected_label"] = opt.get("label","선택값")
    out["selected_basis"] = opt.get("basis","")
    for k in ["best_distance","matched_n","next_vals","n","mean","std","avg_companies","method"]:
        if k in opt:
            out[k] = opt[k]
    return out

def parse_xls(file_bytes, filename=""):
    """입찰서류함 파일 파싱 — xls/xlsx 모두 지원"""
    bids = []
    # ── xlsx 형식 ────────────────────────────────────────────
    is_xlsx = filename.lower().endswith(".xlsx") if filename else False
    if not is_xlsx:
        try:
            # xlsx 매직바이트 확인 (PK = zip 헤더)
            is_xlsx = file_bytes[:2] == b'PK'
        except Exception:
            is_xlsx = False

    if is_xlsx:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows_data = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(rows_data) < 3:
            return bids
        headers = [str(c) if c is not None else "" for c in rows_data[1]]
        for row_vals in rows_data[2:]:
            row = {headers[i]: row_vals[i] for i in range(len(headers))}
            if not row.get("번호"):
                continue
            try:
                base = float(row.get("기초금액") or 0)
            except (ValueError, TypeError):
                base = 0
            bids.append({
                "no":       int(float(row["번호"])),
                "name":     str(row.get("공고명") or ""),
                "bid_no":   str(row.get("공고번호") or ""),
                "base":     base,
                "base_억":  round(base/1e8, 4) if base else 0,
                "deadline": str(row.get("투찰마감") or ""),
                "org":      str(row.get("발주기관") or ""),
                "region":   str(row.get("지역") or ""),
            })
    else:
        # ── xls 형식 (나라장터 기본) ──────────────────────────
        wb = xlrd.open_workbook(file_contents=file_bytes, ignore_workbook_corruption=True)
        ws = wb.sheets()[0]
        headers = [ws.cell_value(1, c) for c in range(ws.ncols)]
        for r in range(2, ws.nrows):
            row = {headers[c]: ws.cell_value(r, c) for c in range(ws.ncols)}
            if not row.get("번호"):
                continue
            base = float(row.get("기초금액") or 0)
            bids.append({
                "no":       int(row["번호"]),
                "name":     row.get("공고명", ""),
                "bid_no":   row.get("공고번호", ""),
                "base":     base,
                "base_억":  round(base/1e8, 4) if base else 0,
                "deadline": row.get("투찰마감", ""),
                "org":      row.get("발주기관", ""),
                "region":   row.get("지역", ""),
            })
    return bids

# ── 영문 변환 ─────────────────────────────────────────────────
def tr_trend(t):
    return {"↑상승":"Up","↓하락":"Down","→횡보":"Flat"}.get(t,t)
def tr_pattern(p):
    return {"연속성":"Momentum","반전":"Reversal","무작위":"Random"}.get(p,p)
def tr_org(org):
    return (org.replace("한국전력공사 ","KEPCO ").replace("본부","")
               .replace("한국철도공사 회계통합센터","KORAIL")
               .replace("조달청","PPS").replace("국군재정관리단","MND"))

# ── 흐름 차트 ─────────────────────────────────────────────────
def make_flow_chart(a1,a2,a3,lo,hi,org_raw):
    all_v=(a1.get("all_vals") or a1.get("recent10",[])) if a1 else []
    if not all_v: return None
    org_en=tr_org(org_raw)
    show_n=min(30,len(all_v)); recent=all_v[-show_n:]
    x=np.arange(1,show_n+1); next_x=show_n+1; mean_v=a1["mean"]

    fig=plt.figure(figsize=(13,5.5),facecolor="#f8fafc")
    gs=fig.add_gridspec(2,1,height_ratios=[2.6,1.0],hspace=0.04)
    ax=fig.add_subplot(gs[0]); ax_l=fig.add_subplot(gs[1])
    ax.set_facecolor("#ffffff"); ax_l.set_facecolor("#f8fafc"); ax_l.axis("off")

    ax.axhline(0,color="#94a3b8",lw=1.0,alpha=0.8,zorder=1)
    if lo is not None and hi is not None:
        ax.axhspan(lo,hi,alpha=0.12,color="#7c3aed",zorder=1)
        ax.axhline(lo,color="#7c3aed",lw=0.7,ls=":",alpha=0.5,zorder=1)
        ax.axhline(hi,color="#7c3aed",lw=0.7,ls=":",alpha=0.5,zorder=1)
    ax.axhline(mean_v,color="#f59e0b",lw=1.3,ls="--",alpha=0.85,zorder=2)
    ma5=[np.mean(recent[max(0,i-4):i+1]) for i in range(show_n)]
    ax.plot(x,ma5,color="#6366f1",lw=1.6,ls="--",alpha=0.75,zorder=3)
    bar_c=["#10b981" if v>=0 else "#ef4444" for v in recent]
    ax.bar(x,recent,color=bar_c,alpha=0.42,width=0.65,zorder=2)
    ax.plot(x,recent,color="#1a2744",lw=1.6,marker="o",ms=3.5,zorder=4)
    for i in range(max(0,show_n-10),show_n):
        v=recent[i]
        ax.annotate(f"{v:+.3f}",xy=(x[i],v),xytext=(0,7 if v>=0 else -11),
                    textcoords="offset points",ha="center",fontsize=6.5,
                    color="#059669" if v>=0 else "#dc2626",fontweight="bold")
    ax.axvline(show_n+0.5,color="#7c3aed",lw=1.4,ls=":",alpha=0.75)

    preds=[]
    if a1: preds.append(("(1)Pattern",a1["pred"],"#1d4ed8","D",10))
    if a2: preds.append(("(2)Similar",a2["pred"],"#15803d","s",10))
    if a3: preds.append(("(3)Trend",  a3["pred"],"#92400e","^",10))
    xoff=[-0.32,0.0,0.32]
    for idx,(lbl,pv,pc,mk,ms) in enumerate(preds):
        px=next_x+xoff[idx]
        ax.plot([px],[pv],color=pc,marker=mk,ms=ms,zorder=7,
                markeredgecolor="white",markeredgewidth=1.1)
        ax.annotate(f"{pv:+.4f}%",xy=(px,pv),xytext=(20,0),
                    textcoords="offset points",ha="left",fontsize=8,color=pc,
                    fontweight="bold",arrowprops=dict(arrowstyle="->",color=pc,lw=1.1))

    if lo is not None and hi is not None:
        bx=next_x+1.3
        ax.annotate("",xy=(bx,lo),xytext=(bx,hi),
                    arrowprops=dict(arrowstyle="<->",color="#7c3aed",lw=2.0))
        ax.text(bx+0.15,(lo+hi)/2,f"Rec.\n{lo:+.4f}\n~{hi:+.4f}",
                fontsize=7.5,color="#7c3aed",fontweight="bold",va="center",ha="left",
                bbox=dict(boxstyle="round,pad=0.3",fc="#f3e8ff",ec="#7c3aed",alpha=0.9))

    ax_r=ax.twinx(); ax_r.set_ylim(ax.get_ylim())
    tks=[mean_v]; tlbls=[f"Avg:{mean_v:+.3f}"]
    if lo is not None: tks+=[lo,hi]; tlbls+=[f"Lo:{lo:+.3f}",f"Hi:{hi:+.3f}"]
    ax_r.set_yticks(tks); ax_r.set_yticklabels(tlbls,fontsize=7,color="#64748b")
    ax.set_xlim(0.3,next_x+2.8)
    ax.set_xticks(list(x)+[next_x])
    ax.set_xticklabels([f"-{show_n-i}" for i in range(show_n)]+["Pred"],fontsize=7)
    ax.set_ylabel("Pred/Base(0%) %",fontsize=8)
    ax.tick_params(labelsize=7.5); ax.grid(axis="y",alpha=0.18,ls="--")
    grade=a1.get("grade","?"); mae=a1.get("mae",0)
    ax.set_title(
        f"{org_en}  |  Last {show_n} results  |  Trend:{tr_trend(a1['trend'])}"
        f"  |  Pattern:{tr_pattern(a1['pattern'])}  |  n={a1['n']}"
        f"  |  Grade:{grade}(MAE:{mae:.3f}%)",
        fontsize=9,fontweight="bold",color="#1a2744",pad=8)

    # 범례 패널
    ax_l.set_xlim(0,1); ax_l.set_ylim(0,1)
    ax_l.add_patch(mpatches.FancyBboxPatch(
        (0.005,0.03),0.990,0.94,boxstyle="round,pad=0.01",
        facecolor="#ffffff",edgecolor="#cbd5e1",linewidth=1.0,
        transform=ax_l.transAxes))
    col_defs=[(0.01,"Chart Legend","#1d4ed8"),(0.265,"Prediction","#15803d"),
              (0.515,"Pattern Detail","#7c3aed"),(0.765,"3-Point Strategy","#991b1b")]
    for cx,htxt,hcol in col_defs:
        ax_l.add_patch(mpatches.FancyBboxPatch(
            (cx+0.002,0.80),0.238,0.16,boxstyle="round,pad=0.005",
            facecolor=hcol,alpha=0.12,edgecolor="none",transform=ax_l.transAxes))
        ax_l.text(cx+0.012,0.885,htxt,fontsize=9,fontweight="bold",
                  color=hcol,va="center",transform=ax_l.transAxes)
    for lx in [0.255,0.505,0.755]:
        ax_l.plot([lx,lx],[0.04,0.97],color="#e2e8f0",lw=1.0,transform=ax_l.transAxes)

    a2n=a2["n"] if a2 else "-"
    grade_color={"A":"#15803d","B":"#1d4ed8","C":"#854d0e","D":"#991b1b"}.get(grade,"#475569")
    items=[
        (0.01,"line","#1a2744","","Actual bid result",  f"Last {show_n} results"),
        (0.01,"line","#6366f1","","Moving Avg MA(5)",    "5-case moving average"),
        (0.01,"line","#f59e0b","","Overall Average",     f"Avg:{mean_v:+.4f}%"),
        (0.01,"band","#7c3aed","","Recommended Zone",    f"{lo:+.4f}%~{hi:+.4f}%" if lo else "-"),
        (0.265,"mark","#1d4ed8","D","(1) Pattern",       f"w={a1['w5']}/{a1['w10']}/{a1['wm']} -> {a1['pred']:+.4f}%"),
        (0.265,"mark","#15803d","s","(2) Similar",       f"n={a2n} -> {a2['pred']:+.4f}%" if a2 else "No data"),
        (0.265,"mark","#92400e","^","(3) Trend",         f"{a3['pred']:+.4f}%" if a3 else "No data"),
        (0.515,"dot",grade_color,"","Accuracy Grade",    f"Grade:{grade} MAE:{mae:.3f}%"),
        (0.515,"dot","#7c3aed","","Last 5 avg (r5)",     f"{a1['r5']:+.4f}%"),
        (0.515,"dot","#7c3aed","","Last 10 avg (r10)",   f"{a1['r10']:+.4f}%"),
        (0.765,"dot","#1d4ed8","","Pattern Scope",       "하단 4개 범위 차트 확인"),
        (0.765,"dot","#15803d","","Similar/Trend",       "보조 판단값"),
        (0.765,"dot","#7c3aed","","Recommended",         "종합 권장구간"),
    ]
    row_cnt={0.01:0,0.265:0,0.515:0,0.765:0}
    TOP_Y=0.72; ROW_GAP=0.225
    for (cx,itype,color,mk,label,desc) in items:
        ri=row_cnt[cx]; row_cnt[cx]+=1; y=TOP_Y-ri*ROW_GAP
        if itype=="line":
            ax_l.plot([cx+0.005,cx+0.038],[y+0.05,y+0.05],color=color,lw=2.4,
                      transform=ax_l.transAxes,clip_on=False)
        elif itype=="band":
            ax_l.add_patch(mpatches.FancyBboxPatch(
                (cx+0.005,y+0.02),0.033,0.065,boxstyle="round,pad=0.003",
                facecolor=color,alpha=0.28,edgecolor=color,linewidth=1.1,
                transform=ax_l.transAxes))
        elif itype in ("mark","dot"):
            ms_val=9 if itype=="mark" else 5.5
            ax_l.plot([cx+0.022],[y+0.05],marker=mk if itype=="mark" else "o",
                      color=color,ms=ms_val,transform=ax_l.transAxes,clip_on=False,
                      markeredgecolor="white" if itype=="mark" else color,
                      markeredgewidth=1.3 if itype=="mark" else 0,alpha=0.75 if itype=="dot" else 1)
        ax_l.text(cx+0.048,y+0.100,label,fontsize=8.5,fontweight="bold",
                  color="#1e293b",va="top",transform=ax_l.transAxes)
        ax_l.text(cx+0.048,y+0.005,desc,fontsize=8,color="#475569",
                  va="top",transform=ax_l.transAxes)
    ax_l.plot([0.01,0.99],[0.045,0.045],color="#e2e8f0",lw=0.8,transform=ax_l.transAxes)
    plt.subplots_adjust(left=0.055,right=0.92,top=0.95,bottom=0.02)
    buf=io.BytesIO()
    plt.savefig(buf,format="png",dpi=140,bbox_inches="tight",facecolor="#f8fafc")
    buf.seek(0); plt.close(); return buf

# ── 한전 세분화 함수 ──────────────────────────────────────────
DIAG_KWS=['광학','초음파','VLF','PD','콘크리트']

def is_kepco(org): return '한국전력공사' in str(org)
def is_diag(name): return any(kw in str(name) for kw in DIAG_KWS)
def is_supervision(name): return '감리' in str(name)

def _sector_vals(df_c, org_filter, name_kws):
    if df_c is None: return np.array([])
    mask = df_c['발주기관'].str.contains(org_filter, na=False)
    kw_mask = pd.Series([False]*len(df_c), index=df_c.index)
    for kw in name_kws: kw_mask = kw_mask | df_c['공고명'].str.contains(kw, na=False)
    return df_c[mask & kw_mask]['예가/기초(0%)'].values

def _sector_stat(vals, scope):
    if len(vals)<3: return None
    n=len(vals); m=float(np.mean(vals)); s=float(np.std(vals))
    r5=float(np.mean(vals[-5:])) if n>=5 else m
    r10=float(np.mean(vals[-10:])) if n>=10 else m
    pred=0.25*r5+0.20*r10+0.55*m
    return {"pred":round(pred,4),"mean":round(m,4),"std":round(s,4),
            "r5":round(r5,4),"r10":round(r10,4),"n":n,
            "all_vals":vals.tolist(),"recent10":[round(float(v),4) for v in vals[-10:]],
            "conservative":round(pred-s*0.4,4),"aggressive":round(pred+s*0.4,4),
            "scope":scope}

def analyze_diag_all(df_c):
    v=_sector_vals(df_c,'한국전력공사','|'.join(DIAG_KWS).split('|'))
    return _sector_stat(v,'All KEPCO Diagnosis')

def analyze_diag_org(org,df_c):
    v=_sector_vals(df_c,org,DIAG_KWS)
    return _sector_stat(v,f"{org.replace('한국전력공사 ','').replace('본부','')} Diagnosis")

def analyze_sup_all(df_c):
    v=_sector_vals(df_c,'한국전력공사',['감리'])
    return _sector_stat(v,'All KEPCO Supervision')

def analyze_sup_org(org,df_c):
    v=_sector_vals(df_c,org,['감리'])
    return _sector_stat(v,f"{org.replace('한국전력공사 ','').replace('본부','')} Supervision")

def _field_keywords(name):
    if is_diag(name):
        return "진단", DIAG_KWS
    if is_supervision(name):
        return "감리", ["감리"]
    return "감리/진단", ["감리"] + DIAG_KWS

def _pattern_vals(df_c, org_filter=None, keywords=None):
    if df_c is None or len(df_c)==0:
        return np.array([])
    mask = pd.Series([True]*len(df_c), index=df_c.index)
    if org_filter:
        mask = mask & df_c['발주기관'].astype(str).str.contains(org_filter, na=False, regex=False)
    if keywords:
        kw_mask = pd.Series([False]*len(df_c), index=df_c.index)
        for kw in keywords:
            kw_mask = kw_mask | df_c['공고명'].astype(str).str.contains(kw, na=False, regex=False)
        mask = mask & kw_mask
    vals = df_c.loc[mask, '예가/기초(0%)'].dropna().values
    return vals

def build_pattern_scopes(org, name, df_c):
    field_label, keywords = _field_keywords(name)
    org_short = str(org).replace('한국전력공사 ','').replace('본부','')
    scopes = [
        _sector_stat(_pattern_vals(df_c, '한국전력공사'), '한국전력공사 전체 패턴'),
        _sector_stat(_pattern_vals(df_c, org), f'{org_short} 전체 패턴'),
        _sector_stat(_pattern_vals(df_c, '한국전력공사', keywords), f'한국전력공사 {field_label} 전체 패턴'),
        _sector_stat(_pattern_vals(df_c, org, keywords), f'{org_short} {field_label} 패턴'),
    ]
    return [s for s in scopes if s]

def make_pattern_scope_chart(scopes, lo, hi, title):
    if not scopes:
        return None
    n = min(4, len(scopes))
    fig, axes = plt.subplots(2, 2, figsize=(13, 7.2), facecolor='#f8fafc')
    axes = axes.flatten()
    colors = ['#1d4ed8', '#dc2626', '#15803d', '#7c3aed']
    for idx, ax in enumerate(axes):
        ax.set_facecolor('#ffffff')
        if idx >= n:
            ax.axis('off')
            continue
        data = scopes[idx]
        vals = data.get('all_vals') or data.get('recent10', [])
        show_n = min(30, len(vals))
        recent = vals[-show_n:]
        x = np.arange(1, show_n+1)
        mv = data['mean']; pv = data['pred']; color = colors[idx]
        ax.axhline(0, color='#94a3b8', lw=1.0, alpha=0.7, zorder=1)
        ax.axhline(mv, color='#f59e0b', lw=1.2, ls='--', alpha=0.8, zorder=2)
        if lo is not None and hi is not None:
            ax.axhspan(lo, hi, alpha=0.10, color='#7c3aed', zorder=1)
        ma5 = [np.mean(recent[max(0, i-4):i+1]) for i in range(show_n)]
        ax.plot(x, ma5, color='#6366f1', lw=1.3, ls='--', alpha=0.7, zorder=3)
        bar_c = [color if v >= 0 else '#94a3b8' for v in recent]
        ax.bar(x, recent, color=bar_c, alpha=0.36, width=0.65, zorder=2)
        ax.plot(x, recent, color='#1a2744', lw=1.2, marker='o', ms=3.0, zorder=4)
        ax.axvline(show_n+0.5, color='#7c3aed', lw=1.0, ls=':', alpha=0.7)
        ax.plot([show_n+1], [pv], marker='D', color='#7c3aed', ms=8, zorder=7,
                markeredgecolor='white', markeredgewidth=1.0)
        ax.annotate(f'{pv:+.4f}%', xy=(show_n+1, pv), xytext=(13,0),
                    textcoords='offset points', ha='left', fontsize=7.5,
                    color='#7c3aed', fontweight='bold',
                    arrowprops=dict(arrowstyle='->', color='#7c3aed', lw=0.9))
        ax.set_xlim(0.3, show_n+3.0)
        ax.set_xticks(list(x[::max(1, show_n//6)]) + [show_n+1])
        ax.tick_params(labelsize=7)
        ax.grid(axis='y', alpha=0.16, ls='--')
        ax.set_title(
            f"{data['scope']} | n={data['n']} | 평균 {mv:+.4f}% | 표준편차 {data['std']:.4f}%",
            fontsize=8.5, fontweight='bold', color='#1a2744', pad=6
        )
    fig.suptitle(title, fontsize=11, fontweight='bold', color='#1a2744', y=0.995)
    plt.tight_layout(pad=1.2)
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=130, bbox_inches='tight', facecolor='#f8fafc')
    buf.seek(0); plt.close(); return buf

def make_sector_chart(d_all,d_org,lo,hi,title):
    datasets=[]
    if d_all: datasets.append((d_all,d_all['scope'],'#1d4ed8'))
    if d_org: datasets.append((d_org,d_org['scope'],'#dc2626'))
    if not datasets: return None
    ncols=len(datasets)
    fig,axes=plt.subplots(1,ncols,figsize=(13,4.5),facecolor='#f8fafc',sharey=False)
    if ncols==1: axes=[axes]
    for ax,(data,scope,color) in zip(axes,datasets):
        ax.set_facecolor('#ffffff')
        vals=data.get('all_vals') or data.get('recent10',[])
        show_n=min(30,len(vals)); recent=vals[-show_n:]
        x=np.arange(1,show_n+1); mv=data['mean']; pv=data['pred']
        ax.axhline(0,color='#94a3b8',lw=1.0,alpha=0.7,zorder=1)
        ax.axhline(mv,color='#f59e0b',lw=1.3,ls='--',alpha=0.8,zorder=2)
        if lo and hi: ax.axhspan(lo,hi,alpha=0.10,color='#7c3aed',zorder=1)
        ma5=[np.mean(recent[max(0,i-4):i+1]) for i in range(show_n)]
        ax.plot(x,ma5,color='#6366f1',lw=1.5,ls='--',alpha=0.7,zorder=3)
        bar_c=[color if v>=0 else '#94a3b8' for v in recent]
        ax.bar(x,recent,color=bar_c,alpha=0.38,width=0.65,zorder=2)
        ax.plot(x,recent,color='#1a2744',lw=1.5,marker='o',ms=3.5,zorder=4)
        for i in range(max(0,show_n-10),show_n):
            v=recent[i]
            ax.annotate(f'{v:+.3f}',xy=(x[i],v),xytext=(0,7 if v>=0 else -11),
                        textcoords='offset points',ha='center',fontsize=6,
                        color='#059669' if v>=0 else '#dc2626',fontweight='bold')
        ax.axvline(show_n+0.5,color='#7c3aed',lw=1.2,ls=':',alpha=0.7)
        ax.plot([show_n+1],[pv],marker='D',color='#7c3aed',ms=9,zorder=7,
                markeredgecolor='white',markeredgewidth=1.2)
        ax.annotate(f'{pv:+.4f}%',xy=(show_n+1,pv),xytext=(16,0),
                    textcoords='offset points',ha='left',fontsize=8,color='#7c3aed',fontweight='bold',
                    arrowprops=dict(arrowstyle='->',color='#7c3aed',lw=1))
        ax.set_xlim(0.3,show_n+3.2)
        ax.set_xticks(list(x)+[show_n+1])
        ax.set_xticklabels([f'-{show_n-i}' for i in range(show_n)]+['Pred'],fontsize=7)
        ax.set_ylabel('Pred/Base(0%) %',fontsize=8)
        ax.tick_params(labelsize=7.5); ax.grid(axis='y',alpha=0.18,ls='--')
        ax.set_title(f'{scope}  (n={data["n"]})',fontsize=9.5,fontweight='bold',color='#1a2744',pad=6)
    fig.suptitle(title,fontsize=10,fontweight='bold',color='#1a2744',y=1.01)
    plt.tight_layout(pad=0.8)
    buf=io.BytesIO()
    plt.savefig(buf,format='png',dpi=130,bbox_inches='tight',facecolor='#f8fafc')
    buf.seek(0); plt.close(); return buf

# ── 엑셀 생성 ─────────────────────────────────────────────────
def make_excel(results):
    from openpyxl import Workbook
    from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
    from openpyxl.utils import get_column_letter
    NAVY="FF1a2744";BLUE="FFdbeafe";GREEN="FFdcfce7";AMBER="FFfef9c3"
    RED_L="FFfee2e2";PURP="FFf3e8ff";GRAY="FFf8fafc"
    thin=Side(style="thin",color="FFd1d5db"); bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
    def H(ws,r,c,v,bg=NAVY,fg="FFFFFFFF",sz=10,bold=True,wrap=False):
        cell=ws.cell(row=r,column=c,value=v)
        cell.font=Font(name="맑은 고딕",bold=bold,color=fg,size=sz)
        cell.fill=PatternFill("solid",start_color=bg)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=wrap)
        cell.border=bdr; return cell
    def C(ws,r,c,v,bg=None,bold=False,right=False,sz=10,color="FF1e293b",center=False,wrap=False):
        cell=ws.cell(row=r,column=c,value=v)
        cell.font=Font(name="맑은 고딕",bold=bold,size=sz,color=color)
        ha="right" if right else ("center" if center else "left")
        cell.alignment=Alignment(horizontal=ha,vertical="center",wrap_text=wrap)
        cell.border=bdr
        if bg: cell.fill=PatternFill("solid",start_color=bg)
        return cell

    wb=Workbook(); ws=wb.active; ws.title="투찰전략"; ws.sheet_view.showGridLines=False
    today=datetime.now().strftime("%Y.%m.%d")
    ws.merge_cells("A1:L1"); t=ws["A1"]
    t.value=f"투찰전략 분석표 — {today}  ★ 선택값 기준"
    t.font=Font(name="맑은 고딕",bold=True,size=13,color="FF1a2744")
    t.fill=PatternFill("solid",start_color="FFe0e7ff")
    t.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=30

    hdrs=["No","공고명","발주기관","기초금액(억)","마감",
          "①패턴(%)","②유사표본(%)","②최소거리","③트렌드(%)","③권장구간","권장하한(%)","권장상한(%)"]
    wids=[5,40,20,10,12,10,10,10,10,14,10,10]
    for i,(h,w) in enumerate(zip(hdrs,wids),1):
        H(ws,2,i,h,wrap=True); ws.column_dimensions[get_column_letter(i)].width=w
    ws.row_dimensions[2].height=36

    for i,row in enumerate(results):
        r=i+3; bg=GRAY if r%2==0 else "FFFFFFFF"
        b=row["bid"]
        a1=row.get("a1_selected") or row["a1"]
        a2=row.get("a2_selected") or row["a2"]
        a3=row.get("a3_selected") or row["a3"]
        lo=row.get("range_lo_selected", row["range_lo"])
        hi=row.get("range_hi_selected", row["range_hi"])
        C(ws,r,1,b["no"],bg=bg,bold=True,center=True)
        C(ws,r,2,b["name"][:48],bg=bg,sz=9,wrap=True)
        C(ws,r,3,b["org"],bg=bg,sz=9)
        if b["base"]>0:
            cx=C(ws,r,4,b["base_억"],bg=bg,right=True); cx.number_format="#,##0.0000"
        else: C(ws,r,4,"미정",bg=bg,center=True,sz=9)
        C(ws,r,5,b["deadline"],bg=bg,sz=9,center=True)
        for ci,a,cbg in [(6,a1,BLUE),(7,a2,GREEN),(9,a3,AMBER)]:
            if a:
                cx2=C(ws,r,ci,a["pred"],bg=cbg,right=True,bold=True,
                      color="FF1d4ed8" if a["pred"]>=0 else "FF991b1b")
                cx2.number_format="+0.0000;-0.0000"
            else: C(ws,r,ci,"이력없음",bg=RED_L,center=True,sz=8)
        if a2 and a2.get("best_distance") is not None:
            cx_d=C(ws,r,8,a2["best_distance"],bg=GREEN,right=True,bold=True,color="FF15803d")
            cx_d.number_format="0.0000"
        else:
            C(ws,r,8,"-",bg=GREEN,center=True,sz=8)
        C(ws,r,10,a3.get("trend_range","-") if a3 else "-",bg=AMBER,center=True,bold=True,color="FF854d0e")
        for ci,val in [(11,lo),(12,hi)]:
            if val is not None:
                cx3=C(ws,r,ci,val,bg=PURP,right=True,bold=True,color="FF7c3aed")
                cx3.number_format="+0.0000;-0.0000"
            else: C(ws,r,ci,"-",bg=bg,center=True)
        ws.row_dimensions[r].height=34

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ════════════════════════════════════════════════════════════════
#  메인 UI
# ════════════════════════════════════════════════════════════════
    st.markdown("""
<div class="main-header">
<h2>📊 투찰전략 분석 시스템</h2>
<p style="margin:0;opacity:0.8">자동추천 + 입찰자 선택형 전략표 | v2.4 | 낙찰이력 기반</p>
</div>""", unsafe_allow_html=True)

with st.sidebar:
    st.header("⚙️ 시스템 설정")
    mode=st.radio("모드 선택",["📊 투찰전략 분석","🔧 배포자 관리"])
    st.divider()
    df_hist=load_history(); pattern_stats=load_pattern_stats()
    if df_hist is not None:
        df_c_s=df_hist[df_hist["예가/기초(0%)"].notna()&(df_hist["예가/기초(0%)"].abs()<10)]
        n_c=len(df_c_s); n_o=df_c_s["발주기관"].nunique()
        st.success(f"✅ 낙찰이력 {n_c:,}건\n{n_o}개 발주처")
    else:
        st.warning("⚠️ 낙찰이력 없음"); df_c_s=None; n_c=0; n_o=0
    if pattern_stats:
        st.success(f"✅ 패턴통계 {len(pattern_stats)}개 발주처")
    st.divider()
    st.caption(f"🕐 {datetime.now().strftime('%Y-%m-%d %H:%M')}")

# ══ 배포자 관리 ══════════════════════════════════════════════════
if mode=="🔧 배포자 관리":
    st.header("🔧 배포자 관리")
    pwd=st.text_input("관리자 비밀번호",type="password")
    ADMIN_PWD=st.secrets.get("ADMIN_PWD",None)
    if not ADMIN_PWD:
        st.error("관리자 비밀번호가 설정되지 않았습니다. Streamlit Secrets에 ADMIN_PWD를 등록하세요.")
        st.stop()
    if pwd!=ADMIN_PWD: st.info("비밀번호를 입력하세요."); st.stop()
    st.success("✅ 관리자 인증")
    uploaded=st.file_uploader("낙찰이력 xlsx 업로드",type=["xlsx","xls"])
    if uploaded:
        with st.spinner("처리 중..."):
            try:
                content=uploaded.read()
                df_new=pd.read_excel(io.BytesIO(content))
                required=["발주기관","공고명","기초금액","예가/기초(0%)"]
                missing=[c for c in required if c not in df_new.columns]
                if missing: st.error(f"필수 컬럼 없음: {missing}")
                else:
                    save_history(df_new)
                    df_v=df_new[df_new["예가/기초(0%)"].notna()&(df_new["예가/기초(0%)"].abs()<10)]
                    st.success("✅ 업로드 완료!")
                    c1,c2,c3=st.columns(3)
                    c1.metric("총 건수",f"{len(df_v):,}건")
                    c2.metric("발주처 수",f"{df_v['발주기관'].nunique()}개")
                    c3.metric("평균 사정율",f"{df_v['예가/기초(0%)'].mean():+.4f}%")
            except Exception as e: st.error(f"오류: {e}")

# ══ 투찰전략 분석 ════════════════════════════════════════════════
else:
    df_hist=load_history()
    df_c = df_hist[df_hist["예가/기초(0%)"].notna()&(df_hist["예가/기초(0%)"].abs()<10)].copy() if df_hist is not None else None
    pattern_stats=load_pattern_stats()

    st.header("📊 투찰전략 분석")
    col_up,col_info=st.columns([2,1])
    with col_up:
        xls_file=st.file_uploader("입찰서류함 xls 파일 업로드",type=["xls","xlsx"])
    with col_info:
        nc=n_c if 'n_c' in dir() else (len(df_c) if df_c is not None else 0)
        no=n_o if 'n_o' in dir() else (df_c['발주기관'].nunique() if df_c is not None else 0)
        st.markdown(f"""
        <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:12px;font-size:0.9em">
        <b>📌 분석 방법</b><br>
        🔵 ①패턴: 한전/동일발주처/분야별 이력 패턴차트<br>
        🟢 ②유사표본: 유사용역 낙찰이력<br>
        🟡 ③트렌드: 최근 흐름 분석<br>
        🟣 💡권장: 3가지 종합 권장구간<br>
        <b>데이터:</b> {nc:,}건 | {no}개 발주처
        </div>""",unsafe_allow_html=True)

    if not xls_file:
        st.info("👆 입찰서류함 xls 파일을 업로드하면 자동 분석합니다."); st.stop()

    with st.spinner("파일 읽는 중..."):
        try:
            raw_bytes = xls_file.read()
            # 낙찰이력 파일 오업로드 감지 (5MB 이상 + xlsx)
            if len(raw_bytes) > 3_000_000 and xls_file.name.lower().endswith(".xlsx"):
                st.error(
                    "⚠️ **잘못된 파일입니다.**\n\n"
                    "- 지금 업로드한 파일: **낙찰이력 데이터** (배포자 관리 전용)\n"
                    "- 여기서 필요한 파일: **나라장터 입찰서류함 xls**\n\n"
                    "👉 낙찰이력 업로드는 **사이드바 → 배포자 관리** 에서 진행하세요."
                )
                st.stop()
            bids = parse_xls(raw_bytes, xls_file.name)
            if not bids:
                st.error(
                    "입찰 건을 읽을 수 없습니다.\n\n"
                    "나라장터에서 다운받은 **입찰서류함 xls** 파일을 업로드해 주세요.\n"
                    "낙찰이력 파일(낙찰데이터.xlsx)은 **배포자 관리** 탭에서 업로드하세요."
                )
                st.stop()
        except Exception as e:
            st.error(
                f"파일 읽기 오류: {e}\n\n"
                "나라장터 입찰서류함 xls 파일인지 확인해 주세요."
            )
            st.stop()

    st.success(f"✅ {len(bids)}건 확인")
    results=[]
    prog=st.progress(0,"분석 중...")
    for i,b in enumerate(bids):
        a1=analyze_pattern(b["org"],df_c,pattern_stats)
        a2=analyze_similar(b["name"],b["base"],df_c,b.get("region",""))
        a3=analyze_trend(b["org"],b["name"],df_c,b.get("region",""))
        lo,hi=recommend_range(a1,a2,a3)
        conv_std,conv_lbl=convergence_score(a1,a2,a3)
        amt_lbl,amt_adj,amt_note=get_amt_info(b["base_억"])
        results.append({"bid":b,"a1":a1,"a2":a2,"a3":a3,
                        "range_lo":lo,"range_hi":hi,
                        "conv_std":conv_std,"conv_lbl":conv_lbl,
                        "amt_lbl":amt_lbl,"amt_adj":amt_adj,"amt_note":amt_note})
        prog.progress((i+1)/len(bids))
    prog.empty()

    # ── 요약 테이블 ──────────────────────────────────────────
    st.subheader(f"📋 자동추천 요약 — {datetime.now().strftime('%Y.%m.%d')} ({len(bids)}건)")
    rows=[]
    for row in results:
        b=row["bid"]; a1=row["a1"]; a2=row["a2"]; a3=row["a3"]
        lo,hi=row["range_lo"],row["range_hi"]
        grade=a1.get("grade","?") if a1 else "?"
        ge={"A":"🟢","B":"🔵","C":"🟡","D":"🔴"}.get(grade,"⚪")
        rows.append({"No":b["no"],
            "공고명":b["name"][:33]+"…" if len(b["name"])>33 else b["name"],
            "발주기관":b["org"].replace("한국전력공사 ","한전 "),
            "기초(억)":f"{b['base_억']:.4f}" if b["base"]>0 else "미정",
            "마감":b["deadline"],
            "①패턴":f"{a1['pred']:+.4f}%" if a1 else "없음",
            "②유사표본":f"{a2['pred']:+.4f}%" if a2 else "없음",
            "②거리":f"{a2['best_distance']:.4f}" if a2 and a2.get("best_distance") is not None else "-",
            "③트렌드":f"{a3['pred']:+.4f}%" if a3 else "없음",
            "③권장":a3.get("trend_range","-") if a3 else "-",
            "💡하한":f"{lo:+.4f}%" if lo else "-",
            "💡상한":f"{hi:+.4f}%" if hi else "-",
            "수렴도":row["conv_lbl"],"등급":f"{ge}{grade}"})
    st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True,
                 column_config={"No":st.column_config.NumberColumn(width=50),
                                "공고명":st.column_config.TextColumn(width=200)})
    st.divider()

    # ── 건별 상세 ────────────────────────────────────────────
    st.subheader("📌 건별 상세 + 사정율 흐름 차트")
    for detail_i,row in enumerate(results):
        b=row["bid"]; a1=row["a1"]; a2=row["a2"]; a3=row["a3"]
        lo,hi=row["range_lo"],row["range_hi"]
        grade=a1.get("grade","?") if a1 else "?"
        ge={"A":"🟢","B":"🔵","C":"🟡","D":"🔴"}.get(grade,"⚪")
        label=(f"No.{b['no']}  {b['name'][:48]}  |  "
               f"{b['org'].replace('한국전력공사 ','한전 ')}  |  "
               f"{b['base_억']:.4f}억  |  {b['deadline']}  {ge}{grade}")
        with st.expander(label):
            # 신뢰도 + 금액구간 배너
            col_g,col_a=st.columns([1,2])
            with col_g:
                gc={"A":"grade-a","B":"grade-b","C":"grade-c","D":"grade-d"}.get(grade,"grade-c")
                mae_v=a1.get("mae",0) if a1 else 0
                st.markdown(f'<span class="{gc}">신뢰도 {grade} (MAE:{mae_v:.3f}%)</span>',unsafe_allow_html=True)
            with col_a:
                amt_note=row["amt_note"]; amt_adj=row["amt_adj"]; amt_lbl=row["amt_lbl"]
                if "유리" in amt_note: st.success(f"💰 {amt_lbl} — 유리한 구간 ({amt_adj:+.4f}%)")
                elif "보수" in amt_note: st.warning(f"⚠️ {amt_lbl} — 보수적 접근 권장 ({amt_adj:+.4f}%)")
                else: st.info(f"📐 {amt_lbl} ({amt_adj:+.4f}%)")

            c1,c2,c3,c4=st.columns(4)
            with c1:
                v=f"{a1['pred']:+.4f}%" if a1 else "이력없음"
                st.markdown(f'<div class="val-box val-pattern">①패턴<br>{v}</div>',unsafe_allow_html=True)
                if a1:
                    st.caption(f"n={a1['n']}건 | {a1['trend']} | {a1['pattern']}패턴")
                    st.caption(f"r5={a1['r5']:+.4f} / r10={a1['r10']:+.4f} / 직전:{a1['last_val']:+.4f}%")
                    if a1.get("std",0) > 0.3:
                        st.caption(f"⚠️ std={a1['std']:.3f}% — 실제분포 넓음")
                pat_opt = select_candidate_ui("① 선택", pattern_candidates(a1), f"pat_{b['no']}_{detail_i}", "기준값(추천)")
            with c2:
                v=f"{a2['pred']:+.4f}%" if a2 else "이력없음"
                fb = a2.get("fallback",False) if a2 else False
                box_style = "val-similar" if not fb else "val-trend"
                st.markdown(f'<div class="val-box {box_style}">②유사표본{"(대체)" if fb else ""}<br>{v}</div>',unsafe_allow_html=True)
                if a2:
                    if fb:
                        st.caption(f"⚠️ {a2.get('fallback_note','분야평균 대체')}")
                    st.caption(f"{a2.get('method','유사표본')} | 유사 {a2['n']}건 | 평균:{a2['mean']:+.4f}%")
                    st.caption(f"성격:{a2.get('service_label','-')} | 지역:{a2.get('region_label') or '전체'} | 업체수구간:{a2.get('company_bucket') or '전체'}")
                    if a2.get("best_distance") is not None:
                        st.caption(f"최소거리 MAE:{a2['best_distance']:.4f} | 유사패턴 {a2.get('matched_n',0)}개 | 다음값 후보:{a2.get('next_vals',[])}")
                    if a2.get("avg_companies"): st.caption(f"업체수 참고:{a2['avg_companies']}개")
                sim_opt = select_candidate_ui("② 선택", a2.get("candidates",[]) if a2 else [], f"sim_{b['no']}_{detail_i}", a2.get("selected_label") if a2 else None)
            with c3:
                v=f"{a3['pred']:+.4f}%" if a3 else "이력없음"
                st.markdown(f'<div class="val-box val-trend">③트렌드<br>{v}</div>',unsafe_allow_html=True)
                if a3:
                    st.caption(f"{a3.get('method','트렌드')} | {a3.get('field_label','전체')} | 표본 {a3.get('scope_n',0)}건")
                    st.caption(f"직전{a3['recent_n']}건 평균:{a3['recent_mean']:+.4f}% | 최근3건:{a3['recent3_mean']:+.4f}%")
                    st.caption(f"다음 권장구간: {a3.get('trend_range','-')} | 유사패턴 {a3.get('matched_n',0)}개")
                    if a3.get("best_distance") is not None:
                        st.caption(f"최소거리 MAE:{a3['best_distance']:.4f} | 다음값 후보:{a3.get('next_vals',[])}")
                trend_opt = select_candidate_ui("③ 선택", a3.get("candidates",[]) if a3 else [], f"trend_{b['no']}_{detail_i}", a3.get("selected_label") if a3 else None)
            sel_a1 = selected_analysis(a1, pat_opt)
            sel_a2 = selected_analysis(a2, sim_opt)
            sel_a3 = selected_analysis(a3, trend_opt)
            lo_sel, hi_sel = recommend_range(sel_a1, sel_a2, sel_a3)
            conv_std_sel, conv_lbl_sel = convergence_score(sel_a1, sel_a2, sel_a3)
            row["a1_selected"] = sel_a1
            row["a2_selected"] = sel_a2
            row["a3_selected"] = sel_a3
            row["range_lo_selected"] = lo_sel
            row["range_hi_selected"] = hi_sel
            row["conv_std_selected"] = conv_std_sel
            row["conv_lbl_selected"] = conv_lbl_sel
            with c4:
                if lo_sel is not None:
                    st.markdown(f'<div class="val-box val-rec">💡선택값 전략<br>{lo_sel:+.4f}%~{hi_sel:+.4f}%</div>',unsafe_allow_html=True)
                    if b["base"]>0:
                        st.caption(f"하한: {int(b['base']*(100+lo_sel)/100):,}원")
                        st.caption(f"상한: {int(b['base']*(100+hi_sel)/100):,}원")
                    st.caption(f"수렴도: {conv_lbl_sel}")
                    st.caption(f"① {sel_a1.get('selected_label','-') if sel_a1 else '-'}")
                    st.caption(f"② {sel_a2.get('selected_label','-') if sel_a2 else '-'}")
                    st.caption(f"③ {sel_a3.get('selected_label','-') if sel_a3 else '-'}")
                else:
                    st.markdown('<div class="val-box" style="background:#fee2e2;color:#991b1b">⚠️ 데이터부족</div>',unsafe_allow_html=True)

            # ── 흐름 차트 ─────────────────────────────────────
            if a1 and (a1.get("all_vals") or a1.get("recent10")):
                st.markdown("---")
                with st.spinner("차트 생성 중..."):
                    chart_buf=make_flow_chart(sel_a1,sel_a2,sel_a3,lo_sel,hi_sel,b["org"])
                if chart_buf: st.image(chart_buf,use_container_width=True)
            else:
                st.caption("⚠️ 이력 데이터 부족")

            # ── ① 패턴 범위 비교 차트 ─────────────────────────
            if df_c is not None:
                st.markdown("---")
                st.markdown("**① 패턴 범위 비교**")
                scopes = build_pattern_scopes(b["org"], b["name"], df_c)
                if scopes:
                    cols = st.columns(min(4, len(scopes)))
                    for col, scope in zip(cols, scopes):
                        with col:
                            st.metric(scope["scope"], f"{scope['pred']:+.4f}%", f"n={scope['n']}건")
                            st.caption(f"평균 {scope['mean']:+.4f}% / 표준편차 {scope['std']:.4f}%")
                    with st.spinner("패턴 범위 비교차트..."):
                        scope_buf = make_pattern_scope_chart(scopes, lo_sel, hi_sel, f"Pattern Scope — {tr_org(b['org'])}")
                    if scope_buf: st.image(scope_buf, use_container_width=True)
                else:
                    st.caption("패턴 범위 비교를 위한 표본이 부족합니다.")

    st.divider()
    st.subheader("✅ 최종 선택값 기준 투찰전략표")
    final_rows=[]
    for row in results:
        b=row["bid"]
        s1=row.get("a1_selected") or row["a1"]
        s2=row.get("a2_selected") or row["a2"]
        s3=row.get("a3_selected") or row["a3"]
        lo=row.get("range_lo_selected", row["range_lo"])
        hi=row.get("range_hi_selected", row["range_hi"])
        conv=row.get("conv_lbl_selected", row["conv_lbl"])
        final_rows.append({
            "No":b["no"],
            "공고명":b["name"][:36]+"…" if len(b["name"])>36 else b["name"],
            "①선택":s1.get("selected_label","-") if s1 else "-",
            "①값":f"{s1['pred']:+.4f}%" if s1 else "-",
            "②선택":s2.get("selected_label","-") if s2 else "-",
            "②값":f"{s2['pred']:+.4f}%" if s2 else "-",
            "③선택":s3.get("selected_label","-") if s3 else "-",
            "③값":f"{s3['pred']:+.4f}%" if s3 else "-",
            "권장하한":f"{lo:+.4f}%" if lo is not None else "-",
            "권장상한":f"{hi:+.4f}%" if hi is not None else "-",
            "수렴도":conv,
        })
    st.dataframe(pd.DataFrame(final_rows),use_container_width=True,hide_index=True)

    st.divider()
    st.subheader("💾 전략표 다운로드")
    excel_buf=make_excel(results)
    today_str=datetime.now().strftime("%Y%m%d")
    st.download_button("📥 엑셀 다운로드 (선택값 기준)",
        data=excel_buf,
        file_name=f"투찰전략_{today_str}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",use_container_width=True)
