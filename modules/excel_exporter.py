from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def make_strategy_excel(results):
    NAVY="FF1a2744"; BLUE="FFdbeafe"; GREEN="FFdcfce7"; AMBER="FFfef9c3"
    RED_L="FFfee2e2"; PURP="FFf3e8ff"; GRAY="FFf8fafc"; RED2="FFfca5a5"; GRN2="FFbbf7d0"
    thin=Side(style="thin",color="FFd1d5db"); bdr=Border(left=thin,right=thin,top=thin,bottom=thin)

    def H(ws,r,c,v,bg=NAVY,fg="FFFFFFFF",sz=10,bold=True,wrap=False):
        cell=ws.cell(row=r,column=c,value=v)
        cell.font=Font(name="맑은 고딕",bold=bold,color=fg,size=sz)
        cell.fill=PatternFill("solid",start_color=bg)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=wrap)
        cell.border=bdr
        return cell

    def C(ws,r,c,v,bg=None,bold=False,right=False,sz=10,color="FF1e293b",center=False,wrap=False):
        cell=ws.cell(row=r,column=c,value=v)
        cell.font=Font(name="맑은 고딕",bold=bold,size=sz,color=color)
        ha="right" if right else ("center" if center else "left")
        cell.alignment=Alignment(horizontal=ha,vertical="center",wrap_text=wrap)
        cell.border=bdr
        if bg:
            cell.fill=PatternFill("solid",start_color=bg)
        return cell

    wb=Workbook(); ws=wb.active; ws.title="투찰전략"; ws.sheet_view.showGridLines=False
    today=datetime.now().strftime("%Y.%m.%d")
    ws.merge_cells("A1:N1")
    t=ws["A1"]
    t.value=f"투찰전략 분석표 — {today}  ★ 3가지 분석값 + 3개업체 분산투찰 전략"
    t.font=Font(name="맑은 고딕",bold=True,size=13,color="FF1a2744")
    t.fill=PatternFill("solid",start_color="FFe0e7ff")
    t.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=30

    hdrs=["No","공고명","발주기관","기초금액(억)","마감","①패턴(%)","②유사표본(%)","③트렌드(%)","권장하한(%)","권장상한(%)","업체A(%)","업체B(%)","업체C(%)","커버율"]
    wids=[5,40,20,10,12,10,10,10,10,10,10,10,10,8]
    for i,(h,w) in enumerate(zip(hdrs,wids),1):
        H(ws,2,i,h,wrap=True); ws.column_dimensions[get_column_letter(i)].width=w
    ws.row_dimensions[2].height=36

    for i,row in enumerate(results):
        r=i+3; bg=GRAY if r%2==0 else "FFFFFFFF"
        b=row["bid"]; a1=row["a1"]; a2=row["a2"]; a3=row["a3"]
        lo,hi=row["range_lo"],row["range_hi"]; tp=row.get("three_pt")
        C(ws,r,1,b["no"],bg=bg,bold=True,center=True)
        C(ws,r,2,b["name"][:60],bg=bg,sz=9,wrap=True)
        C(ws,r,3,b["org"],bg=bg,sz=9)
        if b["base"]>0:
            cx=C(ws,r,4,b["base_억"],bg=bg,right=True); cx.number_format="#,##0.0000"
        else:
            C(ws,r,4,"미정",bg=bg,center=True,sz=9)
        C(ws,r,5,b["deadline"],bg=bg,sz=9,center=True)
        for ci,a,cbg in [(6,a1,BLUE),(7,a2,GREEN),(8,a3,AMBER)]:
            if a:
                cx2=C(ws,r,ci,a["pred"],bg=cbg,right=True,bold=True,color="FF1d4ed8" if a["pred"]>=0 else "FF991b1b")
                cx2.number_format="+0.0000;-0.0000"
            else:
                C(ws,r,ci,"이력없음",bg=RED_L,center=True,sz=8)
        for ci,val in [(9,lo),(10,hi)]:
            if val is not None:
                cx3=C(ws,r,ci,val,bg=PURP,right=True,bold=True,color="FF7c3aed")
                cx3.number_format="+0.0000;-0.0000"
            else:
                C(ws,r,ci,"-",bg=bg,center=True)
        if tp:
            cxa=C(ws,r,11,tp["pt_a"],bg=RED2,right=True,bold=True,color="FF991b1b"); cxa.number_format="+0.00;-0.00"
            cxb=C(ws,r,12,tp["pt_b"],bg=BLUE,right=True,bold=True,color="FF1d4ed8"); cxb.number_format="+0.0000;-0.0000"
            cxc=C(ws,r,13,tp["pt_c"],bg=GRN2,right=True,bold=True,color="FF15803d"); cxc.number_format="+0.00;-0.00"
            C(ws,r,14,f"{tp['cover']}%",bg=PURP,center=True,bold=True,color="FF7c3aed")
        else:
            for ci in [11,12,13,14]: C(ws,r,ci,"-",bg=bg,center=True)
        ws.row_dimensions[r].height=34
    ws.freeze_panes="A3"

    ws2=wb.create_sheet("3포인트 전략"); ws2.sheet_view.showGridLines=False
    ws2.merge_cells("A1:H1")
    t2=ws2["A1"]; t2.value="3개 업체 분산투찰 전략표"
    t2.font=Font(name="맑은 고딕",bold=True,size=12,color="FF1a2744")
    t2.fill=PatternFill("solid",start_color="FFe0e7ff")
    t2.alignment=Alignment(horizontal="center",vertical="center")
    ws2.row_dimensions[1].height=28
    hdrs2=["No","공고명","발주기관","기초금액","업체A포인트","업체B포인트(차트예측)","업체C포인트","커버율"]
    wids2=[5,40,20,12,16,20,16,12]
    for i,(h,w) in enumerate(zip(hdrs2,wids2),1):
        H(ws2,2,i,h,wrap=True); ws2.column_dimensions[get_column_letter(i)].width=w
    for i,row in enumerate(results):
        r=i+3; bg=GRAY if r%2==0 else "FFFFFFFF"
        b=row["bid"]; tp=row.get("three_pt")
        C(ws2,r,1,b["no"],bg=bg,center=True,bold=True)
        C(ws2,r,2,b["name"][:60],bg=bg,sz=9,wrap=True)
        C(ws2,r,3,b["org"],bg=bg,sz=9)
        if b["base"]>0:
            cx=C(ws2,r,4,b["base_억"],bg=bg,right=True); cx.number_format="#,##0.0000억"
        else:
            C(ws2,r,4,"미정",bg=bg,center=True)
        if tp:
            pa,pb,pc=tp["pt_a"],tp["pt_b"],tp["pt_c"]
            if b["base"]>0:
                amt_a=int(b["base"]*(100+pa)/100); amt_b=int(b["base"]*(100+pb)/100); amt_c=int(b["base"]*(100+pc)/100)
                C(ws2,r,5,f"{pa:+.2f}% ({amt_a:,}원)",bg=RED2,bold=True,color="FF991b1b",center=True)
                C(ws2,r,6,f"{pb:+.4f}% ({amt_b:,}원)",bg=BLUE,bold=True,color="FF1d4ed8",center=True)
                C(ws2,r,7,f"{pc:+.2f}% ({amt_c:,}원)",bg=GRN2,bold=True,color="FF15803d",center=True)
            else:
                C(ws2,r,5,f"{pa:+.2f}%",bg=RED2,bold=True,color="FF991b1b",center=True)
                C(ws2,r,6,f"{pb:+.4f}%",bg=BLUE,bold=True,color="FF1d4ed8",center=True)
                C(ws2,r,7,f"{pc:+.2f}%",bg=GRN2,bold=True,color="FF15803d",center=True)
            cover_str=f"{tp['cover']}%"
            if tp['cover_r']>0: cover_str += f"\n최근:{tp['cover_r']}%"
            C(ws2,r,8,cover_str,bg=PURP,bold=True,color="FF7c3aed",center=True,wrap=True)
        else:
            for ci in range(5,9): C(ws2,r,ci,"-",bg=bg,center=True)
        ws2.row_dimensions[r].height=36
    ws2.freeze_panes="A3"

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf
